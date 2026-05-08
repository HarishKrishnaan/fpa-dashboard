"""
Excel report generator for the FP&A dashboard.

The workbook is designed to look like a management reporting package:
- Executive Summary
- P&L Report
- Variance Analysis
- Forecast Detail
- Balance Sheet
- Input Data

Pandas writes the data and openpyxl formats it for readability.
"""

from __future__ import annotations

from io import BytesIO
from typing import Dict

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .finance_engine import pnl_pivot_for_display


HEADER_FILL = "1F4E78"
SUBHEADER_FILL = "D9EAF7"
WHITE = "FFFFFF"
LIGHT_BORDER = "D9D9D9"


def _write_df(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str, index: bool = False) -> None:
    """Write a DataFrame, replacing empty tables with a readable placeholder."""
    if df is None or df.empty:
        df = pd.DataFrame({"Message": ["No data available for this section."]})
    df.to_excel(writer, sheet_name=sheet_name, index=index)


def _format_sheet(ws) -> None:
    """Apply lightweight formatting to one Excel sheet."""
    thin = Side(style="thin", color=LIGHT_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top")

    for cell in ws[1]:
        cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
        width = min(max(len(v) for v in values) + 2, 42)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if "margin" in str(ws.cell(row=1, column=cell.column).value).lower() or "pct" in str(ws.cell(row=1, column=cell.column).value).lower():
                    cell.number_format = "0.0%"
                else:
                    cell.number_format = '$#,##0;[Red]($#,##0);-'


def create_excel_report(
    actuals: pd.DataFrame,
    budget: pd.DataFrame,
    forecast: pd.DataFrame,
    pnl: pd.DataFrame,
    variance: pd.DataFrame,
    balance_sheet: pd.DataFrame,
    balance_summary: pd.DataFrame,
    kpis: Dict[str, float],
) -> bytes:
    """Create an in-memory Excel workbook and return bytes for Streamlit download."""
    output = BytesIO()

    executive_summary = pd.DataFrame(
        [
            {"Metric": "Revenue", "Value": kpis.get("Revenue", 0.0)},
            {"Metric": "Gross Profit", "Value": kpis.get("Gross Profit", 0.0)},
            {"Metric": "Gross Margin", "Value": kpis.get("Gross Margin", 0.0)},
            {"Metric": "Operating Income", "Value": kpis.get("Operating Income", 0.0)},
            {"Metric": "Operating Margin", "Value": kpis.get("Operating Margin", 0.0)},
            {"Metric": "Net Income", "Value": kpis.get("Net Income", 0.0)},
            {"Metric": "Net Margin", "Value": kpis.get("Net Margin", 0.0)},
        ]
    )

    actual_pnl = pnl_pivot_for_display(pnl, "Actual")
    budget_pnl = pnl_pivot_for_display(pnl, "Budget")
    base_forecast_pnl = pnl_pivot_for_display(pnl, "Base Forecast")
    upside_forecast_pnl = pnl_pivot_for_display(pnl, "Upside Forecast")
    downside_forecast_pnl = pnl_pivot_for_display(pnl, "Downside Forecast")

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_df(writer, executive_summary, "Executive Summary")
        _write_df(writer, actual_pnl, "Actual P&L")
        _write_df(writer, budget_pnl, "Budget P&L")
        _write_df(writer, base_forecast_pnl, "Base Forecast P&L")
        _write_df(writer, upside_forecast_pnl, "Upside Forecast P&L")
        _write_df(writer, downside_forecast_pnl, "Downside Forecast P&L")
        _write_df(writer, variance, "Variance Analysis")
        _write_df(writer, balance_summary, "Balance Summary")
        _write_df(writer, balance_sheet, "Balance Sheet Detail")
        _write_df(writer, actuals, "Input Actuals")
        _write_df(writer, budget, "Input Budget")
        _write_df(writer, forecast, "Forecast Detail")

        wb = writer.book
        for ws in wb.worksheets:
            _format_sheet(ws)

        # Add a simple definitions tab to show finance understanding.
        ws = wb.create_sheet("Finance Definitions")
        definitions = [
            ("Budget", "Original annual financial plan used as a baseline."),
            ("Forecast", "Updated expectation based on latest actuals and assumptions."),
            ("Actuals", "Financial results that already happened."),
            ("Variance", "Actual minus Budget."),
            ("Revenue", "Money earned from selling products or services."),
            ("COGS", "Direct costs required to deliver revenue."),
            ("Gross Margin", "Gross Profit divided by Revenue."),
            ("Operating Income", "Gross Profit minus Operating Expenses."),
            ("Net Income", "Profit after operating costs, other income/expense, and taxes."),
            ("Balance Sheet", "Snapshot of Assets, Liabilities, and Equity."),
        ]
        ws.append(["Term", "Meaning"])
        for item in definitions:
            ws.append(list(item))
        _format_sheet(ws)

    output.seek(0)
    return output.getvalue()
